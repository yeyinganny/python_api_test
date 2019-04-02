# -*- coding: utf-8 -*-
# @Time   : 2019/3/11 16:58
# @Author : Tammy
# @Email  : 18585053@qq.com
# @File   : read_write_case.py
''' 2）测试用例的数据存储在Excel中，并编写一个从Excel中读取数据的测试类，包含的函数能够读取测试数据，并且能够写回测试结果，要求有返回值。'''
from openpyxl import load_workbook
from API_7.common.public.input_log import InputLog
from API_7.common.public import data_path
from API_7.common.public.configuration import ReadConfiguration


class Read_Write_Case:
    '''编写一个测试类，用来读取测试数据，并回写测试结果'''
    def __init__(self, file_name, sheet_name):
        # 初始化函数
        self.file_name = file_name
        self.sheet_name = sheet_name
        try:
            self.excel = load_workbook(self.file_name)
            self.sheet = self.excel[self.sheet_name]
        except Exception as e:
            InputLog().error(e)

    def read_data(self, section): # section是配置文件里的片段名，可以根据你的指定来读取测试用例
        '''从Excel里读数据，有返回值'''
        # 从配置文件里控制之下哪些用例
        case_id = ReadConfiguration(data_path.config_path).read_other(section, 'case_id')  # 读配置文件获取case_id的值
        # 将excel中的手机号读出来赋值给请求参数中的mobilephone
        tel = Read_Write_Case(data_path.data_path, 'params').read_param(1, 2)
        test_data = []  # 所有用例存在一个列表中
        for i in range(2, self.sheet.max_row+1):
            case = {}   #  每一行数据存在一个字典里
            case['CaseId']=self.sheet.cell(i, 1).value
            case['Module'] = self.sheet.cell(i, 2).value
            case['Title'] = self.sheet.cell(i, 3).value
            case['Method'] = self.sheet.cell(i, 4).value
            case['Url'] = self.sheet.cell(i, 5).value
            case['Param'] = self.sheet.cell(i, 6).value
            # 字符串中find tel replace tel
            if case['Param'] is not None and case['Param'].find('tel') != -1:
                # 查找tel,如果找到返回位置，没找到返回-1
                InputLog().info('请求参数中tel被替换成手机号{}'.format(tel))
                case['Param'] = case['Param'].replace('tel', str(tel))  # 替换tel,替换后还要赋值回去
                new_tel = int(tel) + 1
                Read_Write_Case(data_path.data_path, 'params').write_back(1, 2, str(new_tel))
            case['sql'] = self.sheet.cell(i, 7).value
            case['ExpectedResult'] = self.sheet.cell(i, 8).value
            test_data.append(case)
        self.excel.close()  # 每次操作完关闭掉
        final_data = []
        if case_id == 'all': # 如果case_id == 'all' 存储所有用例
            final_data = test_data
        else: # 否则存储列表中知道的用例
            for i in case_id: # 遍历case_id里面的值[1,2,3,4]
                final_data.append(test_data[i-1])
        return final_data

    def write_back(self,row,col,result):
        '''写会测试结果到excel中，并保存到当前excel'''
        # 在指定的单元格写入数据
        self.sheet.cell(row,col,result)
        self.excel.save(self.file_name) # 写入数据后要保存工作簿
        self.excel.close() # 每次操作完关闭掉

    def read_param(self,row,col):
        # 获取手机号码
        tel=self.sheet.cell(row,col).value
        return tel


if __name__ == '__main__':
    file_name = 'E:\\Python Project\\test\\API_7\\test_case\\test_api.xlsx'

    # 调用read_param函数
    #print(Read_Write_Case(file_name, 'params').read_param(1,2))
    # 调用read_data函数
    Read_Write_Case(file_name, 'recharge').read_data('RechargeCase')
    # 调用read_param函数
    #print(Read_Write_Case(file_name, 'params').read_param(1,2))
    # 调用write_excel函数
    #Read_Write_Case(file_name, 'login').write_back('abc')



